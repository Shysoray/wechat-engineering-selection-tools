from pathlib import Path
from openpyxl import load_workbook
import json

from fitting_catalog_evidence import (
    CATALOG_SOURCE_NAMES,
    canonicalize_unilok_model,
    catalog_inventory,
    load_catalog_text,
    normalize_model,
    normalized_target_models,
    source_evidence_status,
    split_unilok_source_models,
)


ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = ROOT / "outputs/tube_fitting_audit/Swagelok_VIGOUR_Fujikin_UNILOK_JSK_TK-FUJIKIN_FITOK_SUPERLOK_全品牌全型号对应_修复版.xlsx"
OUTPUT_JS = ROOT / "packageFitting/tubeFittingSupplementalDatabase.js"
AUDIT_OUTPUT = ROOT / "outputs/tube_fitting_audit/tube_fitting_source_evidence_audit.json"
SHEET_NAME = "全品牌全型号对应"

BRAND_COLUMNS = [
    {
        "brand": "Swagelok",
        "model": "Swagelok型号",
        "status": "VIGOUR核查状态",
        "note": "统一人工复核项",
    },
    {
        "brand": "FUJIKIN",
        "model": "FUJIKIN对应型号",
        "status": "FUJIKIN核查状态",
        "note": "统一人工复核项",
    },
    {
        "brand": "UNILOK",
        "model": "UNILOK对应型号",
        "status": "UNILOK核查状态",
        "note": "UNILOK备注",
    },
    {
        "brand": "JSK",
        "model": "JSK对应型号",
        "status": "JSK核查状态",
        "note": "JSK备注",
    },
    {
        "brand": "TK-Fujikin",
        "model": "TK-FUJIKIN对应型号",
        "status": "TK-FUJIKIN核查状态",
        "note": "TK-FUJIKIN备注",
    },
    {
        "brand": "FITOK",
        "model": "FITOK对应型号",
        "status": "FITOK核查状态",
        "note": "FITOK备注",
    },
    {
        "brand": "SUPERLOK",
        "model": "SUPERLOK对应型号",
        "status": "SUPERLOK核查状态",
        "note": "SUPERLOK备注",
    },
]

NO_TARGET_STATUSES = {
    "无目录精确型号",
    "VIGOUR目录无同类/未建立规则",
}

SOURCE_SKIP_STATUSES = {
    "无可靠直接对应",
}

DIMENSION_CONFIRMATION_STATUSES = {
    "规则对应",
    "规则候选",
    "多个目录型号，需确认密封形式",
    "目录未列出精确型号",
}


def value(row, headers, name):
    raw = row[headers[name]]
    if raw is None:
        return ""
    return str(raw).strip()


def intern(pool, text):
    if text not in pool:
        pool[text] = len(pool)
    return pool[text]


def flag_bits(*, dimension_confirmation, no_match, special_unmatched):
    bits = 0
    if dimension_confirmation:
        bits |= 1
    if no_match:
        bits |= 2
    if special_unmatched:
        bits |= 4
    return bits


def compact_rows(rows):
    brands = {}
    products = {}
    statuses = {}
    source_statuses = {}
    notes = {}
    directory_statuses = {}
    evidence_statuses = {}
    source_catalogs = {}
    compact = []
    seen = {}
    usable_sources = {
        (item["brand"], item["sourceModel"])
        for item in rows
        if item.get("vigourModel") and not item.get("noMatch")
    }

    for item in rows:
        source_identity = (item["brand"], item["sourceModel"])
        if item.get("noMatch") and source_identity in usable_sources:
            continue
        identity = (item["brand"], item["sourceModel"], item.get("vigourModel") or "")
        if identity in seen:
            existing = seen[identity]
            existing["dimensionConfirmation"] = bool(
                existing.get("dimensionConfirmation") or item.get("dimensionConfirmation")
            )
            existing["noMatch"] = bool(existing.get("noMatch") or item.get("noMatch"))
            if existing.get("specialFeatureMatched") is False or item.get("specialFeatureMatched") is False:
                existing["specialFeatureMatched"] = False
            continue
        seen[identity] = item

    for item in seen.values():
        compact.append([
            intern(brands, item["brand"]),
            item["sourceModel"],
            item.get("vigourModel") or "",
            intern(products, item["productLabel"]),
            flag_bits(
                dimension_confirmation=item.get("dimensionConfirmation", False),
                no_match=item.get("noMatch", False),
                special_unmatched=item.get("specialFeatureMatched") is False,
            ),
            item.get("sizeCode") or "",
            item.get("variantCode") or "",
            intern(statuses, item.get("statusText") or ""),
            intern(source_statuses, item.get("sourceAuditStatus") or ""),
            intern(notes, item.get("noteText") or ""),
            intern(directory_statuses, item.get("directoryStatus") or ""),
            intern(evidence_statuses, item.get("sourceEvidenceStatus") or "workbook_rule"),
            intern(source_catalogs, item.get("sourceCatalog") or ""),
        ])

    def by_index(pool):
        return [text for text, _ in sorted(pool.items(), key=lambda entry: entry[1])]

    return {
        "b": by_index(brands),
        "p": by_index(products),
        "s": by_index(statuses),
        "a": by_index(source_statuses),
        "n": by_index(notes),
        "d": by_index(directory_statuses),
        "e": by_index(evidence_statuses),
        "c": by_index(source_catalogs),
        "r": compact,
    }


def build_rows():
    workbook = load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)
    sheet = workbook[SHEET_NAME]
    header_row = next(sheet.iter_rows(values_only=True))
    headers = {name: index for index, name in enumerate(header_row) if name}

    rows = []
    excluded = []
    catalog_texts = {brand: load_catalog_text(brand) for brand in CATALOG_SOURCE_NAMES}
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        swagelok_model = value(raw_row, headers, "Swagelok型号")
        product = value(raw_row, headers, "产品类别")
        size_code = value(raw_row, headers, "主管/管外径代码")
        variant_code = value(raw_row, headers, "其他规格代码")
        vigour_model = value(raw_row, headers, "VIGOUR对应型号")
        target_models = normalized_target_models(vigour_model)
        vigour_model = " / ".join(target_models)
        vigour_status = value(raw_row, headers, "VIGOUR核查状态")
        unified_note = value(raw_row, headers, "统一人工复核项")

        if not swagelok_model:
            continue

        vigour_no_match = not vigour_model or vigour_status in NO_TARGET_STATUSES
        target_is_confirmed = not vigour_no_match
        target_is_ambiguous = vigour_status != "精确对应"

        for config in BRAND_COLUMNS:
            raw_source_model = value(raw_row, headers, config["model"])
            if not raw_source_model:
                continue

            if config["brand"] == "UNILOK":
                canonical_source = canonicalize_unilok_model(raw_source_model, size_code)
                source_models = split_unilok_source_models(canonical_source)
            elif config["brand"] == "Swagelok":
                source_models = [normalize_model(raw_source_model)]
            else:
                source_models = [raw_source_model]

            source_status = value(raw_row, headers, config["status"])
            source_note = value(raw_row, headers, config["note"])
            notes = [part for part in [source_note, unified_note] if part]
            note_text = "；".join(dict.fromkeys(notes)) or "普通卡套接头目录匹配；正式替代前需复核尺寸、密封形式和压力等级。"
            no_match = not target_is_confirmed or source_status in SOURCE_SKIP_STATUSES

            status_text = (
                source_status
                if source_status in SOURCE_SKIP_STATUSES
                else "暂无 VIGOUR 精确型号" if no_match
                else source_status or vigour_status or "目录匹配"
            )
            directory_status = vigour_status if config["brand"] == "Swagelok" else source_status or vigour_status

            dimension_confirmation = (
                not no_match
                and (
                    config["brand"] != "Swagelok"
                    or source_status in DIMENSION_CONFIRMATION_STATUSES
                    or target_is_ambiguous
                )
            )

            for source_model in source_models:
                evidence_status = source_evidence_status(
                    config["brand"],
                    source_model,
                    catalog_texts.get(config["brand"]),
                )
                if evidence_status == "unverified":
                    excluded.append({
                        "brand": config["brand"],
                        "sourceModel": source_model,
                        "workbookSourceModel": raw_source_model,
                        "vigourModel": "" if no_match else vigour_model,
                        "reason": "source_model_absent_from_catalog",
                    })
                    continue

                rows.append({
                    "brand": config["brand"],
                    "sourceModel": source_model,
                    "vigourModel": "" if no_match else vigour_model,
                    "productLabel": product,
                    "sizeCode": size_code,
                    "variantCode": variant_code,
                    "statusText": status_text,
                    "sourceAuditStatus": source_status,
                    "noteText": note_text,
                    "directoryStatus": directory_status,
                    "dataSource": "全品牌卡套对照运行库",
                    "dimensionConfirmation": dimension_confirmation,
                    "noMatch": no_match,
                    "specialFeatureMatched": False if target_is_ambiguous or source_status == "规则候选" else True,
                    "sourceEvidenceStatus": evidence_status,
                    "sourceCatalog": CATALOG_SOURCE_NAMES.get(config["brand"], "全品牌卡套对照工作簿"),
                    "searchEligible": True,
                })

    represented = {
        brand: {normalize_model(item["sourceModel"]) for item in rows if item["brand"] == brand}
        for brand in CATALOG_SOURCE_NAMES
    }
    catalog_guards = {}
    for brand in CATALOG_SOURCE_NAMES:
        inventory = catalog_inventory(brand, catalog_texts[brand])
        source_models = inventory if brand == "Swagelok" else {f"{model}-SS" for model in inventory}
        catalog_guards[brand] = sorted(
            model for model in source_models if normalize_model(model) not in represented[brand]
        )

    return rows, excluded, catalog_guards


def js_template_text(text):
    return (
        str(text)
        .replace("\\", "\\\\")
        .replace("`", "\\`")
        .replace("${", "\\${")
        .replace("\t", " ")
        .replace("\r", " ")
        .replace("\n", " ")
    )


def write_js(data, catalog_guards):
    brands = json.dumps(data["b"], ensure_ascii=False, separators=(",", ":"))
    products = json.dumps(data["p"], ensure_ascii=False, separators=(",", ":"))
    statuses = json.dumps(data["s"], ensure_ascii=False, separators=(",", ":"))
    source_statuses = json.dumps(data["a"], ensure_ascii=False, separators=(",", ":"))
    notes = json.dumps(data["n"], ensure_ascii=False, separators=(",", ":"))
    directory_statuses = json.dumps(data["d"], ensure_ascii=False, separators=(",", ":"))
    evidence_statuses = json.dumps(data["e"], ensure_ascii=False, separators=(",", ":"))
    source_catalogs = json.dumps(data["c"], ensure_ascii=False, separators=(",", ":"))
    guard_models = json.dumps(catalog_guards, ensure_ascii=False, separators=(",", ":"))
    row_text = "\n".join(
        "\t".join(js_template_text(cell) for cell in row)
        for row in data["r"]
    )
    OUTPUT_JS.write_text(
        "const {fujikinVlokSupplementalMappings}=require('./fujikinVlokSupplementalMappings')\n"
        + "const brands="
        + brands
        + "\nconst products="
        + products
        + "\nconst statuses="
        + statuses
        + "\nconst sourceStatuses="
        + source_statuses
        + "\nconst notes="
        + notes
        + "\nconst directoryStatuses="
        + directory_statuses
        + "\nconst evidenceStatuses="
        + evidence_statuses
        + "\nconst sourceCatalogs="
        + source_catalogs
        + "\nconst catalogNoMatchModelsByBrand="
        + guard_models
        + "\nconst sourceName='全品牌卡套对照运行库'\n"
        + "const matchStatus='卡套目录匹配'\n"
        + "const noMatchStatus='暂无 VIGOUR 精确型号'\n"
        + "const matchNote='普通卡套接头目录匹配，正式替代前请复核尺寸、密封形式和压力等级。'\n"
        + "const noMatchNote='VIGOUR 目录暂无精确型号，禁止用模糊候选替代。'\n"
        + "const rowText=`"
        + row_text
        + "`\nconst decodedTubeFittingMappings=rowText.split('\\n').filter(Boolean).map((line)=>{"
        + "const row=line.split('\\t');const flags=Number(row[4])||0;const noMatch=Boolean(flags&2);"
        + "const targetModels=(row[2]||'').split(/\\s*\\/\\s*/).filter(Boolean);"
        + "const item={brand:brands[row[0]],sourceModel:row[1],vigourModel:targetModels.join(' / ')||undefined,"
        + "productLabel:products[row[3]],productName:products[row[3]],"
        + "sizeCode:row[5]||undefined,variantCode:row[6]||undefined,"
        + "statusText:statuses[row[7]]||(noMatch?noMatchStatus:matchStatus),"
        + "sourceAuditStatus:sourceStatuses[row[8]]||'',"
        + "noteText:notes[row[9]]||(noMatch?noMatchNote:matchNote),"
        + "directoryStatus:directoryStatuses[row[10]]||(noMatch?noMatchStatus:matchStatus),"
        + "sourceEvidenceStatus:evidenceStatuses[row[11]]||'workbook_rule',sourceCatalog:sourceCatalogs[row[12]]||'',searchEligible:true,"
        + "dataSource:sourceName,selectionEligible:false,"
        + "catalogSupplemental:true,tubeFittingSupplemental:true};"
        + "if(targetModels.length>1)item.vigourModels=targetModels;"
        + "if(flags&1)item.dimensionConfirmation=true;"
        + "if(noMatch)item.noMatch=true;"
        + "if(flags&4){item.specialFeatureMatched=false;item.specialFeatureLabel=item.statusText;item.targetFeatureLabel=item.vigourModel||noMatchStatus}"
        + "return item})\n"
        + "const authoritativeVlokKeys=new Set(fujikinVlokSupplementalMappings.map((item)=>`${item.brand}|${item.sourceModel}|${item.vigourModel||''}`))\n"
        + "const tubeFittingSupplementalMappings=decodedTubeFittingMappings.filter((item)=>!authoritativeVlokKeys.has(`${item.brand}|${item.sourceModel}|${item.vigourModel||''}`))\n"
        + "const suppressedFujikinVlokDuplicateCount=decodedTubeFittingMappings.length-tubeFittingSupplementalMappings.length\n"
        + "module.exports={tubeFittingSupplementalMappings,catalogNoMatchModelsByBrand,suppressedFujikinVlokDuplicateCount}\n",
        encoding="utf-8",
    )


def main():
    rows, excluded, catalog_guards = build_rows()
    data = compact_rows(rows)
    write_js(data, catalog_guards)
    AUDIT_OUTPUT.write_text(json.dumps({
        "generatedRowCount": len(data["r"]),
        "excludedUnverifiedCount": len(excluded),
        "catalogGuardCounts": {brand: len(models) for brand, models in catalog_guards.items()},
        "excludedRows": excluded,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Generated {len(data['r'])} tube fitting supplemental rows -> {OUTPUT_JS}")
    print(f"Excluded unverified source rows: {len(excluded)}")
    print(f"Catalog no-match guards: {sum(len(models) for models in catalog_guards.values())}")
    for brand in data["b"]:
        count = sum(1 for row in data["r"] if data["b"][row[0]] == brand)
        print(f"{brand}: {count}")


if __name__ == "__main__":
    main()
